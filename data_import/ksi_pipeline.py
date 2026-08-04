from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
import json
import os
import re
from typing import TYPE_CHECKING, Iterable

import pandas as pd
from dotenv import load_dotenv

if TYPE_CHECKING:
    from sqlalchemy.engine import Engine


BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "output" / "ksi"
SCHEMA_PATH = Path(__file__).resolve().parent / "sql" / "media_ksi_performance.sql"
TABLE_NAME = "ai_bot_media_ksi_performance"
DEFAULT_EXCEL_CHUNK_SIZE = 20_000
DEFAULT_MYSQL_CHUNK_SIZE = 500

load_dotenv(BASE_DIR / ".env")


SOURCE_COLUMN_MAP = {
    "Platform": "platform",
    "Brand": "brand",
    "Tier": "tier",
    "BigV": "big_v_cost",
    "BET": "bet",
    "NickName": "nickname",
    "KOLID(前台)": "kol_id_front",
    "跳转平台": "redirect_platform",
    "Contentformat": "content_format",
    "PublishedAt": "published_at",
    "MONTH": "source_month",
    "followerCount": "follower_count",
    "ViewCount": "view_count",
    "CommentCount": "comment_count",
    "LikeCount": "like_count",
    "RepostCount": "repost_count",
    "CollectCount": "collect_count",
    "TTLengagement": "ttl_engagement",
    "Bigcategory": "big_category",
    "sub_category": "sub_category",
    "SKU": "sku",
    "KOLID(后台)": "kol_id_back",
    "内容id": "content_id",
    "Title": "title",
    "KOLtype": "kol_type",
    "Selectivity": "selectivity",
    "Tier__2": "tier_secondary",
}

CANONICAL_SOURCE_COLUMNS = list(SOURCE_COLUMN_MAP.values())
TEXT_COLUMNS = [
    "platform",
    "brand",
    "tier",
    "nickname",
    "kol_id_front",
    "redirect_platform",
    "content_format",
    "big_category",
    "sub_category",
    "sku",
    "kol_id_back",
    "content_id",
    "title",
    "kol_type",
    "selectivity",
    "tier_secondary",
]
INTEGER_COLUMNS = [
    "source_month",
    "follower_count",
    "view_count",
    "comment_count",
    "like_count",
    "repost_count",
    "collect_count",
    "ttl_engagement",
]
REQUIRED_COLUMNS = [
    "platform",
    "brand",
    "tier",
    "big_v_cost",
    "published_at",
    "kol_type",
]
OUTPUT_COLUMNS = [
    "record_key",
    "period_month",
    "year",
    "month",
    "platform",
    "brand",
    "tier",
    "big_v_cost",
    "bet",
    "nickname",
    "kol_id_front",
    "redirect_platform",
    "content_format",
    "published_at",
    "source_month",
    "follower_count",
    "view_count",
    "comment_count",
    "like_count",
    "repost_count",
    "collect_count",
    "ttl_engagement",
    "big_category",
    "sub_category",
    "sku",
    "kol_id_back",
    "content_id",
    "title",
    "kol_type",
    "selectivity",
    "tier_secondary",
    "source_file",
    "source_sheet",
    "source_row_number",
    "import_batch_id",
]


@dataclass
class SourceQuality:
    source_file: str
    source_sheet: str
    output_file: str
    source_rows: int = 0
    clean_rows: int = 0
    invalid_rows: int = 0
    tier_mismatch_rows: int = 0
    source_month_mismatch_rows: int = 0
    negative_cost_rows: int = 0
    null_ttl_engagement_rows: int = 0
    invalid_by_field: dict[str, int] = field(default_factory=dict)
    total_big_v_cost: float = 0.0
    total_ttl_engagement: int = 0
    partitions: dict[str, dict] = field(default_factory=dict)
    missing_optional_columns: list[str] = field(default_factory=list)
    output_size_bytes: int = 0
    warnings: list[str] = field(default_factory=list)

    def update(self, frame: pd.DataFrame, stats: dict) -> None:
        self.source_rows += stats["source_rows"]
        self.clean_rows += len(frame)
        self.invalid_rows += stats["invalid_rows"]
        for column, count in stats["invalid_by_field"].items():
            self.invalid_by_field[column] = (
                self.invalid_by_field.get(column, 0) + int(count)
            )
        self.tier_mismatch_rows += stats["tier_mismatch_rows"]
        self.source_month_mismatch_rows += stats["source_month_mismatch_rows"]
        self.negative_cost_rows += int((frame["big_v_cost"] < 0).sum())
        self.null_ttl_engagement_rows += int(frame["ttl_engagement"].isna().sum())
        self.total_big_v_cost += float(frame["big_v_cost"].sum())
        self.total_ttl_engagement += int(frame["ttl_engagement"].fillna(0).sum())
        grouped = (
            frame.groupby(["year", "month"], dropna=False)
            .agg(
                rows=("record_key", "size"),
                big_v_cost=("big_v_cost", "sum"),
                ttl_engagement=("ttl_engagement", "sum"),
            )
            .reset_index()
        )
        for _, row in grouped.iterrows():
            key = f"{int(row['year']):04d}-{int(row['month']):02d}"
            target = self.partitions.setdefault(
                key,
                {"year": int(row["year"]), "month": int(row["month"]), "rows": 0,
                 "big_v_cost": 0.0, "ttl_engagement": 0},
            )
            target["rows"] += int(row["rows"])
            target["big_v_cost"] += float(row["big_v_cost"])
            target["ttl_engagement"] += int(row["ttl_engagement"] or 0)

    def as_dict(self) -> dict:
        warnings = list(self.warnings)
        if self.invalid_rows:
            warnings.append(f"跳过 {self.invalid_rows} 行必填字段无效记录")
        if self.tier_mismatch_rows:
            warnings.append(
                f"主 Tier 与第二个 Tier 有 {self.tier_mismatch_rows} 行不一致"
            )
        if self.source_month_mismatch_rows:
            warnings.append(
                f"MONTH 与 PublishedAt 月份有 "
                f"{self.source_month_mismatch_rows} 行不一致"
            )
        if self.negative_cost_rows:
            warnings.append(f"保留 {self.negative_cost_rows} 行负数 Big V")
        return {
            "source_file": self.source_file,
            "source_sheet": self.source_sheet,
            "output_file": self.output_file,
            "source_rows": self.source_rows,
            "clean_rows": self.clean_rows,
            "invalid_rows": self.invalid_rows,
            "invalid_by_field": self.invalid_by_field,
            "missing_optional_columns": self.missing_optional_columns,
            "tier_mismatch_rows": self.tier_mismatch_rows,
            "source_month_mismatch_rows": self.source_month_mismatch_rows,
            "negative_cost_rows": self.negative_cost_rows,
            "null_ttl_engagement_rows": self.null_ttl_engagement_rows,
            "total_big_v_cost": round(self.total_big_v_cost, 6),
            "total_ttl_engagement": self.total_ttl_engagement,
            "partitions": [
                {
                    **value,
                    "big_v_cost": round(value["big_v_cost"], 6),
                }
                for _, value in sorted(self.partitions.items())
            ],
            "output_size_bytes": self.output_size_bytes,
            "warnings": warnings,
        }


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value).replace("\u00a0", " ")).strip()


def unique_headers(values: Iterable[object]) -> list[str]:
    counts: dict[str, int] = {}
    headers: list[str] = []
    for value in values:
        base = normalize_header(value)
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return headers


def normalize_label(value: object) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()
    return text or None


def source_to_canonical(frame: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    canonical = pd.DataFrame(index=frame.index)
    missing_optional: list[str] = []
    for source_name, target_name in SOURCE_COLUMN_MAP.items():
        if source_name in frame.columns:
            canonical[target_name] = frame[source_name]
        else:
            canonical[target_name] = pd.NA
            missing_optional.append(target_name)
    return canonical, missing_optional


def make_record_key(source_file: str, source_sheet: str, row_number: int) -> str:
    payload = f"{source_file}|{source_sheet}|{row_number}"
    return sha256(payload.encode("utf-8")).hexdigest()


def transform_chunk(
    source: pd.DataFrame,
    *,
    source_file: str,
    source_sheet: str,
    batch_id: str,
) -> tuple[pd.DataFrame, dict, list[str]]:
    frame, missing_optional = source_to_canonical(source)
    for column in TEXT_COLUMNS:
        frame[column] = frame[column].map(normalize_label)
    frame["big_v_cost"] = pd.to_numeric(
        frame["big_v_cost"], errors="coerce"
    ).astype("Float64")
    frame["bet"] = pd.to_numeric(frame["bet"], errors="coerce").astype("Float64")
    for column in INTEGER_COLUMNS:
        frame[column] = pd.to_numeric(frame[column], errors="coerce").astype("Int64")
    frame["published_at"] = pd.to_datetime(frame["published_at"], errors="coerce")

    invalid_by_field = {
        column: int(frame[column].isna().sum())
        for column in REQUIRED_COLUMNS
    }
    invalid = pd.Series(False, index=frame.index)
    for column in REQUIRED_COLUMNS:
        invalid |= frame[column].isna()
    stats = {
        "source_rows": int(len(frame)),
        "invalid_rows": int(invalid.sum()),
        "invalid_by_field": invalid_by_field,
        "tier_mismatch_rows": 0,
        "source_month_mismatch_rows": 0,
    }
    frame = frame.loc[~invalid].copy()
    source_rows = pd.to_numeric(
        source.loc[frame.index, "_source_row_number"], errors="raise"
    ).astype(int)

    frame["year"] = frame["published_at"].dt.year.astype(int)
    frame["month"] = frame["published_at"].dt.month.astype(int)
    frame["period_month"] = frame["published_at"].dt.to_period("M").dt.to_timestamp()

    both_tiers = frame["tier"].notna() & frame["tier_secondary"].notna()
    stats["tier_mismatch_rows"] = int(
        (
            both_tiers
            & (
                frame["tier"].str.casefold()
                != frame["tier_secondary"].str.casefold()
            )
        ).sum()
    )
    has_source_month = frame["source_month"].notna()
    stats["source_month_mismatch_rows"] = int(
        (has_source_month & (frame["source_month"] != frame["month"])).sum()
    )

    frame.insert(
        0,
        "record_key",
        [
            make_record_key(source_file, source_sheet, int(row_number))
            for row_number in source_rows
        ],
    )
    frame["source_file"] = source_file
    frame["source_sheet"] = source_sheet
    frame["source_row_number"] = source_rows.to_numpy()
    frame["import_batch_id"] = batch_id
    return frame[OUTPUT_COLUMNS], stats, missing_optional


def iter_excel_chunks(
    input_path: str | Path,
    *,
    chunk_size: int = DEFAULT_EXCEL_CHUNK_SIZE,
):
    from openpyxl import load_workbook

    path = Path(input_path).expanduser().resolve()
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = unique_headers(next(rows))
    buffer: list[tuple] = []
    first_data_row = 2
    for row_number, row in enumerate(rows, start=2):
        buffer.append(tuple(row))
        if len(buffer) >= chunk_size:
            frame = pd.DataFrame(buffer, columns=headers)
            frame["_source_row_number"] = range(
                first_data_row, first_data_row + len(frame)
            )
            yield sheet.title, frame
            buffer = []
            first_data_row = row_number + 1
    if buffer:
        frame = pd.DataFrame(buffer, columns=headers)
        frame["_source_row_number"] = range(
            first_data_row, first_data_row + len(frame)
        )
        yield sheet.title, frame
    workbook.close()


def safe_output_stem(path: Path) -> str:
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", path.stem).strip("_").lower()
    return normalized or "ksi"


def prepare_excel_file(
    input_path: str | Path,
    output_dir: str | Path,
    *,
    batch_id: str,
    chunk_size: int = DEFAULT_EXCEL_CHUNK_SIZE,
) -> tuple[Path, SourceQuality, pd.DataFrame]:
    import pyarrow as pa
    import pyarrow.parquet as pq

    path = Path(input_path).expanduser().resolve()
    directory = Path(output_dir).expanduser().resolve()
    directory.mkdir(parents=True, exist_ok=True)
    output_path = directory / f"media_ksi_{safe_output_stem(path)}.parquet"
    writer = None
    quality: SourceQuality | None = None
    sample_parts: list[pd.DataFrame] = []

    try:
        for sheet_name, source_chunk in iter_excel_chunks(
            path, chunk_size=chunk_size
        ):
            cleaned, stats, missing_optional = transform_chunk(
                source_chunk,
                source_file=path.name,
                source_sheet=sheet_name,
                batch_id=batch_id,
            )
            if quality is None:
                quality = SourceQuality(
                    source_file=path.name,
                    source_sheet=sheet_name,
                    output_file=output_path.name,
                    missing_optional_columns=sorted(set(missing_optional)),
                )
            quality.update(cleaned, stats)
            if sum(len(part) for part in sample_parts) < 1000:
                sample_parts.append(cleaned.head(1000 - sum(len(p) for p in sample_parts)))
            table = pa.Table.from_pandas(cleaned, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(
                    output_path,
                    table.schema,
                    compression="zstd",
                    use_dictionary=True,
                )
            else:
                table = table.cast(writer.schema, safe=False)
            writer.write_table(table)
    finally:
        if writer is not None:
            writer.close()

    if quality is None:
        raise ValueError(f"{path.name}: 没有可处理的数据")
    quality.output_size_bytes = output_path.stat().st_size
    sample = pd.concat(sample_parts, ignore_index=True) if sample_parts else pd.DataFrame()
    return output_path, quality, sample


def prepare_excel(
    input_path: str,
    output_dir: str | Path,
    *,
    batch_id: str | None = None,
    chunk_size: int = DEFAULT_EXCEL_CHUNK_SIZE,
) -> tuple[Path, Path, Path]:
    batch_id = batch_id or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    directory = Path(output_dir).expanduser().resolve()
    directory.mkdir(parents=True, exist_ok=True)
    output_path, quality, sample = prepare_excel_file(
        input_path,
        directory,
        batch_id=batch_id,
        chunk_size=chunk_size,
    )
    stem = safe_output_stem(Path(input_path))
    quality_path = directory / f"media_ksi_{stem}_quality_report.json"
    quality_path.write_text(
        json.dumps(
            {"batch_id": batch_id, "source": quality.as_dict()},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    sample_path = directory / f"media_ksi_{stem}_sample.csv"
    sample.to_csv(sample_path, index=False, encoding="utf-8-sig")
    print(
        f"prepared {quality.source_file}: "
        f"{quality.clean_rows:,} rows -> {output_path.name}",
        flush=True,
    )
    return output_path, quality_path, sample_path


def get_engine() -> "Engine":
    from sqlalchemy import create_engine
    from sqlalchemy.engine import URL

    url = URL.create(
        "mysql+pymysql",
        username=os.environ["MYSQL_USER"],
        password=os.environ["MYSQL_PASSWORD"],
        host=os.environ["MYSQL_HOST"],
        port=int(os.environ.get("MYSQL_PORT", "3306")),
        database=os.environ["MYSQL_DATABASE"],
        query={"charset": "utf8mb4"},
    )
    return create_engine(url, pool_pre_ping=True)


def create_table(engine: "Engine") -> None:
    from sqlalchemy import text

    with engine.begin() as connection:
        connection.execute(text(SCHEMA_PATH.read_text(encoding="utf-8")))


def iter_parquet_frames(
    parquet_paths: list[str | Path],
    *,
    batch_size: int,
    columns: list[str] | None = None,
):
    import pyarrow.parquet as pq

    for parquet_path in parquet_paths:
        parquet = pq.ParquetFile(Path(parquet_path).expanduser().resolve())
        for batch in parquet.iter_batches(batch_size=batch_size, columns=columns):
            yield batch.to_pandas()


def parquet_partitions(parquet_paths: list[str | Path]) -> set[tuple[int, int]]:
    partitions: set[tuple[int, int]] = set()
    for frame in iter_parquet_frames(
        parquet_paths,
        batch_size=50_000,
        columns=["year", "month"],
    ):
        partitions.update(
            (int(year), int(month))
            for year, month in frame[["year", "month"]]
            .drop_duplicates()
            .itertuples(index=False, name=None)
        )
    return partitions


def serialize_records(frame: pd.DataFrame) -> list[dict]:
    records = frame.copy()
    records["period_month"] = pd.to_datetime(records["period_month"]).dt.date
    records["published_at"] = pd.to_datetime(records["published_at"])
    records = records.astype(object).where(pd.notnull(records), None)
    return records.to_dict(orient="records")


def load_parquet_to_mysql(
    parquet_paths: list[str | Path],
    engine: "Engine",
    *,
    chunk_size: int = DEFAULT_MYSQL_CHUNK_SIZE,
) -> int:
    from sqlalchemy import MetaData, Table

    create_table(engine)
    table = Table(TABLE_NAME, MetaData(), autoload_with=engine)
    partitions = parquet_partitions(parquet_paths)
    total = 0
    with engine.begin() as connection:
        for year, month in sorted(partitions):
            connection.execute(
                table.delete().where(
                    (table.c.year == year) & (table.c.month == month)
                )
            )
        for frame in iter_parquet_frames(
            parquet_paths,
            batch_size=chunk_size,
        ):
            rows = serialize_records(frame)
            connection.execute(table.insert(), rows)
            total += len(rows)
            if total % 10_000 < chunk_size:
                print(f"MySQL inserted: {total:,}")
    return total


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="KSI 大文件本地清洗为 Parquet，并分块写入 MySQL。"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare = subparsers.add_parser("prepare", help="Excel 清洗为压缩 Parquet")
    prepare.add_argument(
        "--input",
        required=True,
        help="单个 KSI Excel；大年度文件必须分两个命令处理以控制内存",
    )
    prepare.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
    )
    prepare.add_argument("--batch-id")
    prepare.add_argument(
        "--chunk-size",
        type=int,
        default=DEFAULT_EXCEL_CHUNK_SIZE,
        help="本地 Excel 流式清洗行数",
    )

    load = subparsers.add_parser("load", help="压缩 Parquet 分块写入 MySQL")
    load.add_argument(
        "--parquet",
        nargs="+",
        required=True,
        help="一个或多个标准化 KSI Parquet",
    )
    load.add_argument(
        "--chunk-size",
        type=int,
        default=DEFAULT_MYSQL_CHUNK_SIZE,
        help="每批 MySQL 写入行数",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.command == "prepare":
        path, quality_path, sample_path = prepare_excel(
            args.input,
            args.output_dir,
            batch_id=args.batch_id,
            chunk_size=args.chunk_size,
        )
        print(f"parquet: {path}")
        print(f"quality report: {quality_path}")
        print(f"sample csv: {sample_path}")
        return
    if args.command == "load":
        inserted = load_parquet_to_mysql(
            args.parquet,
            get_engine(),
            chunk_size=args.chunk_size,
        )
        print(f"MySQL partition replace completed: {inserted:,} rows")


if __name__ == "__main__":
    main()
